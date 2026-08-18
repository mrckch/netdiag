"""XLSX-Export der Messungen, sortiert nach Etage → Raum → Dose.

Wählbar: Spalten, letzte-Messung-pro-Dose vs. Historie, Zeitraum.
"""
import datetime
import tempfile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.db import get_conn

# key -> (Spaltenüberschrift, SQL-Ausdruck)
COLUMNS = {
    "floor": ("Etage", "f.name"),
    "room": ("Raum", "r.name"),
    "outlet": ("Dose", "o.label"),
    "device": ("Gerät", "dt.name"),
    "timestamp": ("Zeitpunkt", "m.started_at"),
    "kind": ("Messart", "m.kind"),
    "speed": ("Speed", "m.speed"),
    "duplex": ("Duplex", "m.duplex"),
    "vlan_ids": ("VLAN-IDs", "m.vlan_ids"),
    "switch_name": ("Switch", "m.switch_name"),
    "switch_port": ("Switch-Port", "m.switch_port"),
    "dhcp_ok": ("DHCP", "m.dhcp_ok"),
    "iperf_mbps": ("iperf Mbit/s", "m.iperf_mbps"),
    "notes": ("Notizen", "o.notes"),
}

DEFAULT_COLUMNS = ["floor", "room", "outlet", "device", "timestamp",
                   "speed", "duplex", "vlan_ids", "switch_name", "switch_port", "dhcp_ok"]


def export_xlsx(columns: list[str], latest_only: bool = True,
                date_from: int | None = None, date_to: int | None = None) -> str:
    """Erzeugt die XLSX-Datei und gibt den Pfad zurück."""
    cols = [c for c in columns if c in COLUMNS] or DEFAULT_COLUMNS

    select_parts = [f"{COLUMNS[c][1]} AS {c}" for c in cols]
    where = ["m.outlet_id IS NOT NULL"]
    params: list = []
    if date_from:
        where.append("m.started_at >= ?")
        params.append(date_from)
    if date_to:
        where.append("m.started_at <= ?")
        params.append(date_to)

    latest_filter = ""
    if latest_only:
        latest_filter = """AND m.id = (
            SELECT m2.id FROM measurements m2
            WHERE m2.outlet_id = m.outlet_id
            ORDER BY m2.started_at DESC, m2.id DESC LIMIT 1
        )"""

    sql = f"""
        SELECT {', '.join(select_parts)}
        FROM measurements m
        JOIN outlets o ON o.id = m.outlet_id
        JOIN rooms r ON r.id = o.room_id
        JOIN floors f ON f.id = r.floor_id
        LEFT JOIN device_types dt ON dt.id = o.device_type_id
        WHERE {' AND '.join(where)} {latest_filter}
        ORDER BY f.sort_order, f.name, r.name, o.label, m.started_at
    """
    conn = get_conn()
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Kabelkataster"

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    headers = [COLUMNS[c][0] for c in cols]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font

    for row in rows:
        values = []
        for c in cols:
            v = row[c]
            if c == "timestamp" and v:
                v = datetime.datetime.fromtimestamp(v).strftime("%d.%m.%Y %H:%M")
            elif c == "dhcp_ok" and v is not None:
                v = "ok" if v else "keine Antwort"
            values.append(v)
        ws.append(values)

    for r in ws.iter_rows(min_row=2):
        for cell in r:
            cell.font = body_font

    # Spaltenbreiten grob nach Inhalt
    for i, c in enumerate(cols, start=1):
        max_len = max(
            [len(str(COLUMNS[c][0]))] + [len(str(row[c] or "")) for row in rows]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    path = tempfile.mktemp(suffix=".xlsx", prefix="netdiag-export-")
    wb.save(path)
    return path
